from __future__ import annotations

import datetime as dt
import os
import re
import unicodedata
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

try:
    from win32com.client import Dispatch
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Thiếu pywin32 / win32com.client. Hãy cài pywin32 trước khi chạy tool."
    ) from exc


XL_FILEFORMAT_XLSX = 51


@dataclass
class ProjectData:
    project: str = ""
    location: str = ""
    owner: str = ""
    owner_rep: str = ""
    designer: str = ""
    consultant: str = ""
    contractor: str = ""
    contractor_rep: str = ""


@dataclass
class BatchRow:
    index: int
    raw_number: str
    doc_number: str
    work_type: str
    structure_name: str
    date_text: str
    link: str = ""


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text).strip().lower()


def safe_sheet_name(name: str, existing: Iterable[str]) -> str:
    cleaned = re.sub(r"[\[\]\*\?:/\\]", " ", name)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        cleaned = "Sheet"
    cleaned = cleaned[:31].rstrip()
    existing_set = set(existing)
    candidate = cleaned
    suffix = 1
    while candidate in existing_set:
        base = cleaned[: 31 - len(f"_{suffix}")]
        candidate = f"{base}_{suffix}"
        suffix += 1
    return candidate


def format_vn_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, dt.datetime):
        date_value = value
    elif isinstance(value, dt.date):
        date_value = dt.datetime.combine(value, dt.time())
    else:
        text = str(value).strip()
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                date_value = dt.datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        else:
            return text
    return f"{date_value.day:02d} tháng {date_value.month:02d} năm {date_value.year}"


def to_2d(value: Any) -> list[list[Any]]:
    if value is None:
        return []
    if not isinstance(value, tuple):
        return [[value]]
    first = value[0] if value else ()
    if isinstance(first, tuple):
        return [list(row) for row in value]
    return [list(value)]


def cell_str(cell_value: Any) -> str:
    return "" if cell_value is None else str(cell_value).strip()


def get_sheet_values(ws, cell_refs: dict[str, str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for key, ref in cell_refs.items():
        result[key] = cell_str(ws.Range(ref).Value)
    return result


def read_project_data(ws) -> ProjectData:
    values = get_sheet_values(
        ws,
        {
            "project": "C4",
            "location": "C5",
            "owner": "C6",
            "owner_rep": "C7",
            "designer": "C8",
            "consultant": "C9",
            "contractor": "C11",
            "contractor_rep": "C12",
        },
    )
    return ProjectData(**values)


def parse_batch_rows(ws) -> list[BatchRow]:
    used = ws.UsedRange
    data = to_2d(used.Value2)
    if not data:
        return []
    headers = [normalize_text(h) for h in data[0]]
    col_index = {name: idx for idx, name in enumerate(headers) if name}

    def get(row: list[Any], *names: str) -> str:
        for name in names:
            idx = col_index.get(normalize_text(name))
            if idx is not None and idx < len(row):
                return cell_str(row[idx])
        return ""

    rows: list[BatchRow] = []
    last_number = ""
    suffix = 1
    for excel_row_idx, row in enumerate(data[1:], start=2):
        raw_number = get(row, "SỐ BBNT", "SO BBNT")
        work_type = get(row, "CÔNG TÁC NT", "CONG TAC NT")
        structure_name = get(row, "TÊN CẤU KIỆN", "TEN CAU KIEN")
        date_text = get(row, "NGÀY NGHIỆM THU", "NGAY NGHIEM THU")
        link = get(row, "LINK")

        if not any([raw_number, work_type, structure_name, date_text, link]):
            continue

        if raw_number:
            doc_number = raw_number
            last_number = raw_number
            suffix = 1
        elif last_number:
            suffix += 1
            doc_number = f"{last_number}-{suffix:02d}"
        else:
            doc_number = f"BBNT-{len(rows)+1:03d}"

        rows.append(
            BatchRow(
                index=excel_row_idx,
                raw_number=raw_number,
                doc_number=doc_number,
                work_type=work_type,
                structure_name=structure_name,
                date_text=date_text,
                link=link,
            )
        )
    return rows


def convert_source_to_xlsx(source_path: str) -> tuple[str, str | None]:
    path = Path(source_path)
    if path.suffix.lower() == ".xlsx":
        return str(path.resolve()), None
    temp_dir = Path(tempfile.mkdtemp(prefix="bienban_nt_"))
    converted = temp_dir / f"{path.stem}.xlsx"
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True)
        wb.SaveAs(str(converted), FileFormat=XL_FILEFORMAT_XLSX)
    finally:
        if wb is not None:
            wb.Close(False)
        excel.Quit()
    return str(converted), str(temp_dir)


def read_project_data_xl(ws) -> ProjectData:
    return ProjectData(
        project=cell_str(ws["C4"].value),
        location=cell_str(ws["C5"].value),
        owner=cell_str(ws["C6"].value),
        owner_rep=cell_str(ws["C7"].value),
        designer=cell_str(ws["C8"].value),
        consultant=cell_str(ws["C9"].value),
        contractor=cell_str(ws["C11"].value),
        contractor_rep=cell_str(ws["C12"].value),
    )


def parse_batch_rows_xl(ws) -> list[BatchRow]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(v) for v in rows[0]]
    col_index = {name: idx for idx, name in enumerate(headers) if name}

    def get(row: tuple[Any, ...], *names: str) -> str:
        for name in names:
            idx = col_index.get(normalize_text(name))
            if idx is not None and idx < len(row):
                return cell_str(row[idx])
        return ""

    result: list[BatchRow] = []
    last_number = ""
    suffix = 1
    for excel_row_idx, row in enumerate(rows[1:], start=2):
        raw_number = get(row, "SỐ BBNT", "SO BBNT")
        work_type = get(row, "CÔNG TÁC NT", "CONG TAC NT")
        structure_name = get(row, "TÊN CẤU KIỆN", "TEN CAU KIEN")
        date_text = get(row, "NGÀY NGHIỆM THU", "NGAY NGHIEM THU")
        link = get(row, "LINK")
        if not any([raw_number, work_type, structure_name, date_text, link]):
            continue
        if raw_number:
            doc_number = raw_number
            last_number = raw_number
            suffix = 1
        elif last_number:
            suffix += 1
            doc_number = f"{last_number}-{suffix:02d}"
        else:
            doc_number = f"BBNT-{len(result)+1:03d}"
        result.append(
            BatchRow(
                index=excel_row_idx,
                raw_number=raw_number,
                doc_number=doc_number,
                work_type=work_type,
                structure_name=structure_name,
                date_text=date_text,
                link=link,
            )
        )
    return result


def clear_dynamic_fields_xl(ws) -> None:
    for address in [
        "A2",
        "A6",
        "A7",
        "D8",
        "C8",
        "C11",
        "C14",
        "D17",
        "B18",
        "D19",
        "B20",
        "D20",
        "B21",
        "D21",
        "B22",
        "D22",
        "D23",
        "B24",
        "A28",
        "A29",
    ]:
        ws[address] = ""


def apply_row_data_xl(
    ws,
    row: BatchRow,
    project: ProjectData,
    default_start_time: str,
    default_end_time: str,
) -> None:
    if row.doc_number:
        ws["A2"] = f"Số / No.: {row.doc_number}"
    if project.project:
        ws["A6"] = f"Dự án          : {project.project}"
    if project.location:
        ws["A7"] = f"Địa điểm     : {project.location}"
    if row.structure_name:
        ws["C8"] = f": {row.structure_name}"
        ws["C11"] = row.structure_name
    if row.work_type:
        ws["C14"] = row.work_type
    if row.date_text:
        vn_date = format_vn_date(row.date_text)
        if vn_date:
            ws["A28"] = f"  Bắt đầu : {default_start_time} ngày {vn_date}"
            ws["A29"] = f"  Kết thúc : {default_end_time} ngày {vn_date}"


def write_summary_sheet_xl(ws, batch_rows: list[BatchRow]) -> None:
    ws.append(["STT", "Số biên bản", "Công tác", "Cấu kiện / hạng mục", "Ngày nghiệm thu", "Sheet"])
    for idx, row in enumerate(batch_rows, start=1):
        ws.append([idx, row.doc_number, row.work_type, row.structure_name, row.date_text, "Mở"])


def write_catalog_sheet_xl(ws, batch_rows: list[BatchRow]) -> None:
    ws.append(["SỐ BBNT", "CÔNG TÁC NT", "TÊN CẤU KIỆN", "NGÀY NGHIỆM THU", "LINK"])
    for row in batch_rows:
        ws.append([row.doc_number, row.work_type, row.structure_name, row.date_text, row.link])


def find_template_sheet_name(sheet_names: list[str]) -> str | None:
    for name in sheet_names:
        if normalize_text(name).startswith("nt "):
            return name
    for name in sheet_names:
        if normalize_text(name).startswith("nt"):
            return name
    return None


def set_cell_value(ws, address: str, value: Any) -> None:
    ws.Range(address).Value = value


def clear_dynamic_fields(ws) -> None:
    for address in [
        "A2",
        "A6",
        "A7",
        "D8",
        "C8",
        "C11",
        "C14",
        "D17",
        "B18",
        "D19",
        "B20",
        "D20",
        "B21",
        "D21",
        "B22",
        "D22",
        "D23",
        "B24",
        "A28",
        "A29",
    ]:
        ws.Range(address).Value = ""


def apply_project_data(ws, project: ProjectData) -> None:
    if project.project:
        set_cell_value(ws, "A6", f"Dự án          : {project.project}")
    if project.location:
        set_cell_value(ws, "A7", f"Địa điểm     : {project.location}")
    if project.owner:
        set_cell_value(ws, "D17", f": {project.owner}")
    if project.owner_rep:
        set_cell_value(ws, "B18", f"Ông : {project.owner_rep}")
    if project.designer:
        set_cell_value(ws, "D8", f": {project.designer}")
    if project.consultant:
        set_cell_value(ws, "D19", f": {project.consultant}")
    if project.contractor:
        set_cell_value(ws, "D23", f": {project.contractor}")
    if project.contractor_rep:
        set_cell_value(ws, "B24", f"Ông : {project.contractor_rep}")


def apply_row_data(
    ws,
    row: BatchRow,
    project: ProjectData,
    default_start_time: str,
    default_end_time: str,
) -> None:
    if row.doc_number:
        set_cell_value(ws, "A2", f"Số / No.: {row.doc_number}")
    if project.project:
        set_cell_value(ws, "A6", f"Dự án          : {project.project}")
    if project.location:
        set_cell_value(ws, "A7", f"Địa điểm     : {project.location}")
    if row.structure_name:
        set_cell_value(ws, "C8", f": {row.structure_name}")
    if row.structure_name:
        set_cell_value(ws, "C11", row.structure_name)
    if row.work_type:
        set_cell_value(ws, "C14", row.work_type)
    if row.date_text:
        vn_date = format_vn_date(row.date_text)
        if vn_date:
            set_cell_value(ws, "A28", f"  Bắt đầu : {default_start_time} ngày {vn_date}")
            set_cell_value(ws, "A29", f"  Kết thúc : {default_end_time} ngày {vn_date}")


def write_summary_sheet(ws, batch_rows: list[BatchRow]) -> None:
    headers = [["STT", "Số biên bản", "Công tác", "Cấu kiện / hạng mục", "Ngày nghiệm thu", "Sheet"]]
    ws.Range("A1:F1").Value = headers
    for idx, row in enumerate(batch_rows, start=2):
        ws.Cells(idx, 1).Value = idx - 1
        ws.Cells(idx, 2).Value = row.doc_number
        ws.Cells(idx, 3).Value = row.work_type
        ws.Cells(idx, 4).Value = row.structure_name
        ws.Cells(idx, 5).Value = row.date_text
        ws.Cells(idx, 6).Value = "Mở"


def write_catalog_sheet(ws, batch_rows: list[BatchRow]) -> None:
    headers = [["SỐ BBNT", "CÔNG TÁC NT", "TÊN CẤU KIỆN", "NGÀY NGHIỆM THU", "LINK"]]
    ws.Range("A1:E1").Value = headers
    for idx, row in enumerate(batch_rows, start=2):
        ws.Cells(idx, 1).Value = row.doc_number
        ws.Cells(idx, 2).Value = row.work_type
        ws.Cells(idx, 3).Value = row.structure_name
        ws.Cells(idx, 4).Value = row.date_text
        ws.Cells(idx, 5).Value = row.link


def load_workbook_metadata(path: str) -> tuple[list[str], list[str], str | None]:
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(path, ReadOnly=True)
        sheet_names = [ws.Name for ws in wb.Worksheets]
        template = find_template_sheet_name(sheet_names)
        return sheet_names, sheet_names, template
    finally:
        if wb is not None:
            wb.Close(False)
        excel.Quit()


def generate_output(
    source_path: str,
    output_path: str,
    template_sheet_name: str | None = None,
    batch_sheet_name: str = "DANH MUC NT TSX",
    data_sheet_name: str = "Data",
    default_start_time: str = "15h00'",
    default_end_time: str = "16h30'",
) -> list[BatchRow]:
    source_path = str(Path(source_path).resolve())
    output_path = str(Path(output_path).resolve())
    os.makedirs(Path(output_path).parent, exist_ok=True)
    if os.path.exists(output_path):
        os.remove(output_path)

    converted_source, temp_dir = convert_source_to_xlsx(source_path)
    try:
        wb_data = load_workbook(converted_source, read_only=True, data_only=False)
        sheet_names = wb_data.sheetnames
        if not template_sheet_name:
            template_sheet_name = find_template_sheet_name(sheet_names)
        if not template_sheet_name:
            raise RuntimeError("Không tìm thấy sheet mẫu bắt đầu bằng 'NT'.")
        if batch_sheet_name not in sheet_names:
            raise RuntimeError(f"Không tìm thấy sheet danh mục: {batch_sheet_name}")
        if data_sheet_name not in sheet_names:
            raise RuntimeError(f"Không tìm thấy sheet dữ liệu chung: {data_sheet_name}")

        batch_rows = parse_batch_rows_xl(wb_data[batch_sheet_name])
        project = read_project_data_xl(wb_data[data_sheet_name])
        wb_data.close()
        if not batch_rows:
            raise RuntimeError("Sheet danh mục không có dòng dữ liệu hợp lệ.")

        excel = Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        wb = None
        try:
            wb = excel.Workbooks.Open(converted_source, ReadOnly=False)
            wb.SaveAs(output_path, FileFormat=XL_FILEFORMAT_XLSX)

            # Keep only the sheets we need as source material.
            for sheet in list(wb.Worksheets):
                if sheet.Name not in {template_sheet_name, data_sheet_name, batch_sheet_name}:
                    sheet.Delete()

            summary_ws = wb.Worksheets.Add(Before=wb.Worksheets(1))
            summary_ws.Name = "Tổng hợp"
            write_summary_sheet_com(summary_ws, batch_rows)

            template_ws = wb.Worksheets(template_sheet_name)
            generated_sheet_names: list[str] = []
            for row in batch_rows:
                template_ws.Copy(None, wb.Worksheets(wb.Worksheets.Count))
                ws = wb.Worksheets(wb.Worksheets.Count)
                clear_dynamic_fields(ws)
                apply_row_data(ws, row, project, default_start_time, default_end_time)
                sheet_name = safe_sheet_name(row.doc_number, generated_sheet_names)
                ws.Name = sheet_name
                generated_sheet_names.append(sheet_name)
                row.link = sheet_name

            catalog_ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
            catalog_ws.Name = "Danh mục"
            write_catalog_sheet_com(catalog_ws, batch_rows)

            for idx, row in enumerate(batch_rows, start=2):
                summary_ws.Hyperlinks.Add(
                    Anchor=summary_ws.Cells(idx, 6),
                    Address="",
                    SubAddress=f"'{row.link}'!A1",
                    TextToDisplay="Mở",
                )
                catalog_ws.Hyperlinks.Add(
                    Anchor=catalog_ws.Cells(idx, 5),
                    Address="",
                    SubAddress=f"'{row.link}'!A1",
                    TextToDisplay=row.link,
                )

            # Remove the master template from the final workbook after all clones exist.
            template_ws.Delete()

            for ws in wb.Worksheets:
                try:
                    ws.Rows(1).Font.Bold = True
                    ws.Columns.AutoFit()
                except Exception:
                    pass

            wb.Save()
            return batch_rows
        finally:
            if wb is not None:
                wb.Close(SaveChanges=False)
            excel.Quit()
    finally:
        if temp_dir:
            try:
                for child in Path(temp_dir).glob("*"):
                    child.unlink(missing_ok=True)
                Path(temp_dir).rmdir()
            except Exception:
                pass


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Tool tạo biên bản nghiệm thu hàng loạt")
        self.geometry("980x760")
        self.configure(bg="#0f172a")

        self.source_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.template_sheet = tk.StringVar()
        self.batch_sheet = tk.StringVar(value="DANH MUC NT TSX")
        self.data_sheet = tk.StringVar(value="Data")
        self.start_time = tk.StringVar(value="15h00'")
        self.end_time = tk.StringVar(value="16h30'")
        self.status = tk.StringVar(value="Chưa nạp file")
        self.sheet_options: list[str] = []

        self._build_ui()

    def _build_ui(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TFrame", background="#0f172a")
        style.configure("TLabel", background="#0f172a", foreground="#e2e8f0")
        style.configure("TButton", padding=8)
        style.configure("TEntry", fieldbackground="#111827", foreground="#e2e8f0")
        style.configure("Treeview", background="#111827", fieldbackground="#111827", foreground="#e2e8f0")
        style.configure("Treeview.Heading", background="#1f2937", foreground="#f8fafc")

        root = ttk.Frame(self, padding=16)
        root.pack(fill="both", expand=True)

        title = tk.Label(
            root,
            text="Tool tạo biên bản nghiệm thu hàng loạt",
            bg="#0f172a",
            fg="#f8fafc",
            font=("Segoe UI", 20, "bold"),
        )
        title.pack(anchor="w")

        subtitle = tk.Label(
            root,
            text="Đọc workbook mẫu, chọn sheet template, rồi sinh workbook tổng hợp với các biên bản đã đổ dữ liệu.",
            bg="#0f172a",
            fg="#94a3b8",
            font=("Segoe UI", 10),
        )
        subtitle.pack(anchor="w", pady=(4, 14))

        form = ttk.Frame(root)
        form.pack(fill="x")

        self._row_input(form, 0, "Workbook nguồn", self.source_path, self._browse_source)
        self._row_input(form, 1, "File đầu ra", self.output_path, self._browse_output)
        self._row_select(form, 2, "Sheet template", self.template_sheet)
        self._row_entry(form, 3, "Sheet danh mục", self.batch_sheet)
        self._row_entry(form, 4, "Sheet dữ liệu chung", self.data_sheet)
        self._row_entry(form, 5, "Giờ bắt đầu", self.start_time)
        self._row_entry(form, 6, "Giờ kết thúc", self.end_time)

        actions = ttk.Frame(root)
        actions.pack(fill="x", pady=(14, 10))
        ttk.Button(actions, text="Nạp workbook", command=self.load_metadata).pack(side="left")
        ttk.Button(actions, text="Sinh biên bản", command=self.run_generation).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="Mở thư mục đầu ra", command=self.open_output_folder).pack(side="left", padx=(8, 0))
        ttk.Label(actions, textvariable=self.status).pack(side="right")

        preview_frame = ttk.Frame(root)
        preview_frame.pack(fill="both", expand=True, pady=(8, 0))
        ttk.Label(preview_frame, text="Xem trước sheet").pack(anchor="w")
        self.preview = ttk.Treeview(
            preview_frame,
            columns=("name", "type"),
            show="headings",
            height=10,
        )
        self.preview.heading("name", text="Sheet")
        self.preview.heading("type", text="Vai trò")
        self.preview.column("name", width=250)
        self.preview.column("type", width=120)
        self.preview.pack(fill="x", pady=(6, 14))

        ttk.Label(preview_frame, text="Nhật ký").pack(anchor="w")
        self.log = tk.Text(
            preview_frame,
            height=16,
            bg="#111827",
            fg="#e5e7eb",
            insertbackground="#e5e7eb",
            relief="flat",
            wrap="word",
        )
        self.log.pack(fill="both", expand=True, pady=(6, 0))

    def _row_input(self, parent: ttk.Frame, row: int, label: str, variable: tk.StringVar, browse_cmd) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=4, padx=(0, 10))
        entry = ttk.Entry(parent, textvariable=variable, width=80)
        entry.grid(row=row, column=1, sticky="ew", pady=4)
        ttk.Button(parent, text="Chọn", command=browse_cmd).grid(row=row, column=2, padx=(8, 0))
        parent.grid_columnconfigure(1, weight=1)

    def _row_select(self, parent: ttk.Frame, row: int, label: str, variable: tk.StringVar) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=4, padx=(0, 10))
        combo = ttk.Combobox(parent, textvariable=variable, values=self.sheet_options, width=78, state="readonly")
        combo.grid(row=row, column=1, sticky="ew", pady=4)
        ttk.Label(parent, text="Tự nhận dạng từ workbook", foreground="#64748b").grid(row=row, column=2, sticky="w", padx=(8, 0))
        self.template_combo = combo

    def _row_entry(self, parent: ttk.Frame, row: int, label: str, variable: tk.StringVar) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=4, padx=(0, 10))
        ttk.Entry(parent, textvariable=variable, width=80).grid(row=row, column=1, sticky="ew", pady=4, columnspan=2)

    def browse_file(self, save: bool = False) -> str:
        if save:
            return filedialog.asksaveasfilename(
                title="Chọn file đầu ra",
                defaultextension=".xlsx",
                filetypes=[("Excel workbook", "*.xlsx")],
            )
        return filedialog.askopenfilename(
            title="Chọn workbook nguồn",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")],
        )

    def _browse_source(self) -> None:
        path = self.browse_file(save=False)
        if path:
            self.source_path.set(path)
            self.load_metadata()

    def _browse_output(self) -> None:
        path = self.browse_file(save=True)
        if path:
            self.output_path.set(path)

    def log_line(self, text: str) -> None:
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.update_idletasks()

    def load_metadata(self) -> None:
        path = self.source_path.get().strip()
        if not path:
            messagebox.showwarning("Thiếu file", "Hãy chọn workbook nguồn trước.")
            return
        if not os.path.exists(path):
            messagebox.showerror("Không tồn tại", f"Không tìm thấy file:\n{path}")
            return
        try:
            excel = Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(path, ReadOnly=True)
            try:
                self.sheet_options = [ws.Name for ws in wb.Worksheets]
                self.template_combo["values"] = self.sheet_options
                default_template = find_template_sheet_name(self.sheet_options)
                if default_template:
                    self.template_sheet.set(default_template)
                elif self.sheet_options:
                    self.template_sheet.set(self.sheet_options[0])

                try:
                    batch_ws = wb.Worksheets(self.batch_sheet.get())
                    rows = parse_batch_rows(batch_ws)
                except Exception:
                    rows = []
                self.status.set(f"Đã nạp {len(rows)} dòng danh mục")
                self.preview.delete(*self.preview.get_children())
                for name in self.sheet_options:
                    role = "template" if normalize_text(name).startswith("nt") else "sheet"
                    if name in {self.batch_sheet.get(), self.data_sheet.get()}:
                        role = "data"
                    self.preview.insert("", "end", values=(name, role))
                self.log_line(f"Đã nạp workbook: {path}")
                self.log_line(f"Tìm thấy {len(self.sheet_options)} sheet.")
                self.log_line(f"Sheet template mặc định: {self.template_sheet.get()}")
                self.log_line(f"Dòng danh mục hợp lệ: {len(rows)}")
            finally:
                wb.Close(False)
                excel.Quit()
        except Exception as exc:
            self.status.set("Lỗi khi nạp file")
            messagebox.showerror("Lỗi", str(exc))

    def run_generation(self) -> None:
        source = self.source_path.get().strip()
        output = self.output_path.get().strip()
        if not source:
            messagebox.showwarning("Thiếu file", "Hãy chọn workbook nguồn.")
            return
        if not output:
            messagebox.showwarning("Thiếu file", "Hãy chọn file đầu ra.")
            return
        try:
            self.status.set("Đang sinh biên bản...")
            self.log_line("Bắt đầu sinh workbook đầu ra...")
            rows = generate_output(
                source_path=source,
                output_path=output,
                template_sheet_name=self.template_sheet.get().strip() or None,
                batch_sheet_name=self.batch_sheet.get().strip(),
                data_sheet_name=self.data_sheet.get().strip(),
                default_start_time=self.start_time.get().strip(),
                default_end_time=self.end_time.get().strip(),
            )
            self.status.set(f"Hoàn thành {len(rows)} biên bản")
            self.log_line(f"Đã tạo xong: {output}")
            self.log_line(f"Tổng số biên bản sinh ra: {len(rows)}")
            messagebox.showinfo("Thành công", f"Đã sinh xong {len(rows)} biên bản.\nFile đầu ra:\n{output}")
        except Exception as exc:
            self.status.set("Sinh biên bản thất bại")
            messagebox.showerror("Lỗi sinh biên bản", str(exc))
            self.log_line(f"Lỗi: {exc}")

    def open_output_folder(self) -> None:
        output = self.output_path.get().strip()
        if not output:
            messagebox.showwarning("Thiếu file", "Chưa có đường dẫn đầu ra.")
            return
        folder = Path(output).parent
        os.startfile(folder)


def main() -> None:
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
