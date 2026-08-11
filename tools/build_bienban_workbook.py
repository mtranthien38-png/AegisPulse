from __future__ import annotations

import datetime as dt
import os
import re
import shutil
import sys
import tempfile
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from win32com.client import Dispatch


SOURCE_XLS = Path(r"C:\Users\Asus\Desktop\BIÊN BẢN NT TLS, TCNL.xls")
OUTPUT_XLSM = Path(r"D:\AIRDROP\GENLAYER\MR\AegisPulse\dist\BienBanNghiemThuTool.xlsm")
DESKTOP_COPY = Path(r"C:\Users\Asus\Desktop\BienBanNghiemThuTool.xlsm")


@dataclass
class BatchRow:
    index: int
    doc_number: str
    work_type: str
    structure_name: str
    date_text: str
    link: str = ""


@dataclass
class ProjectData:
    project: str = ""
    location: str = ""
    owner: str = ""
    owner_rep: str = ""
    consultant: str = ""
    contractor: str = ""
    contractor_rep: str = ""


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text).strip().lower()


def cell_str(value: object) -> str:
    return "" if value is None else str(value).strip()


def format_vn_date(value: object) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, dt.datetime):
        d = value
    elif isinstance(value, dt.date):
        d = dt.datetime.combine(value, dt.time())
    else:
        text = str(value).strip()
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                d = dt.datetime.strptime(text, fmt)
                break
            except ValueError:
                d = None
        if d is None:
            return text
    return f"{d.day:02d} tháng {d.month:02d} năm {d.year}"


def safe_sheet_name(name: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\*\?:/\\]", " ", name)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        cleaned = "Sheet"
    cleaned = cleaned[:31].rstrip()
    candidate = cleaned
    suffix = 1
    while candidate in existing:
        suffix += 1
        suffix_text = f"-{suffix:02d}"
        candidate = f"{cleaned[:31 - len(suffix_text)]}{suffix_text}"
    return candidate


def load_source_data(xlsx_path: str) -> tuple[list[BatchRow], ProjectData]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        if "DanhMuc" in wb.sheetnames:
            ws_dm = wb["DanhMuc"]
        else:
            ws_dm = wb["DANH MUC NT TSX"]
        ws_data = wb["ThongTinChung"] if "ThongTinChung" in wb.sheetnames else wb["Data"]

        project = ProjectData(
            project=cell_str(ws_data["C4"].value),
            location=cell_str(ws_data["C5"].value),
            owner=cell_str(ws_data["C6"].value),
            owner_rep=cell_str(ws_data["C7"].value),
            consultant=cell_str(ws_data["C9"].value),
            contractor=cell_str(ws_data["C11"].value),
            contractor_rep=cell_str(ws_data["C12"].value),
        )

        rows = list(ws_dm.iter_rows(values_only=True))
        headers = [normalize_text(v) for v in rows[0]]
        col_index = {name: idx for idx, name in enumerate(headers) if name}

        def get(row: tuple[object, ...], *names: str) -> str:
            for name in names:
                idx = col_index.get(normalize_text(name))
                if idx is not None and idx < len(row):
                    return cell_str(row[idx])
            return ""

        batch_rows: list[BatchRow] = []
        last_number = ""
        suffix = 1
        for excel_row_idx, row in enumerate(rows[1:], start=2):
            raw_number = get(row, "SỐ BBNT", "SO BBNT")
            work_type = get(row, "CÔNG TÁC NT", "CONG TAC NT")
            structure_name = get(row, "TÊN CẤU KIỆN", "TEN CAU KIEN")
            date_text = get(row, "NGÀY NGHIỆM THU", "NGAY NGHIEM THU")
            link = get(row, "LINK")
            if not any([raw_number, work_type, structure_name, date_text, link]):
                continue
            if raw_number:
                doc_number = raw_number
                last_number = raw_number
                suffix = 1
            elif last_number:
                suffix += 1
                doc_number = f"{last_number}-{suffix:02d}"
            else:
                doc_number = f"BBNT-{len(batch_rows)+1:03d}"
            batch_rows.append(BatchRow(excel_row_idx, doc_number, work_type, structure_name, date_text, link))
        return batch_rows, project
    finally:
        wb.close()


def convert_source_to_xlsx(source_path: Path) -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix="bbnt_source_"))
    out_path = temp_dir / f"{source_path.stem}.xlsx"
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(source_path), ReadOnly=True)
        wb.SaveAs(str(out_path), FileFormat=51)
        return out_path
    finally:
        if wb is not None:
            wb.Close(False)
        excel.Quit()


VBA_MODULE = r'''
Option Explicit

Private Const SHEET_QM As String = "QuanLy"
Private Const SHEET_DM As String = "DanhMuc"
Private Const SHEET_DATA As String = "ThongTinChung"
Private Const SHEET_TEMPLATE As String = "Mau_NT"

Public Sub TaoBienBanHangLoat()
    Dim wsDM As Worksheet, wsData As Worksheet, wsTemplate As Worksheet, wsNew As Worksheet, wsQL As Worksheet
    Dim lastRow As Long, r As Long, sheetName As String, generatedCount As Long
    Dim usedNames As Object

    Set usedNames = CreateObject("Scripting.Dictionary")

    On Error GoTo CleanFail
    Application.ScreenUpdating = False
    Application.DisplayAlerts = False
    Application.EnableEvents = False

    Set wsDM = ThisWorkbook.Worksheets(SHEET_DM)
    Set wsData = ThisWorkbook.Worksheets(SHEET_DATA)
    Set wsTemplate = ThisWorkbook.Worksheets(SHEET_TEMPLATE)
    Set wsQL = ThisWorkbook.Worksheets(SHEET_QM)

    DeleteGeneratedSheets
    PrepareDashboard wsQL
    PrepareCatalog wsDM
    PrepareDataSheet wsData
    EnsureActionButtons wsQL

    lastRow = Application.WorksheetFunction.Max( _
        wsDM.Cells(wsDM.Rows.Count, "A").End(xlUp).Row, _
        wsDM.Cells(wsDM.Rows.Count, "B").End(xlUp).Row, _
        wsDM.Cells(wsDM.Rows.Count, "C").End(xlUp).Row, _
        wsDM.Cells(wsDM.Rows.Count, "D").End(xlUp).Row)

    For r = 2 To lastRow
        If RowHasData(wsDM, r) Then
            sheetName = UniqueSheetName(RowSheetBase(wsDM, r), usedNames)
            wsTemplate.Copy After:=ThisWorkbook.Worksheets(ThisWorkbook.Worksheets.Count)
            Set wsNew = ActiveSheet
            wsNew.Name = sheetName
            FillAcceptanceSheet wsNew, wsData, wsDM, r
            FormatGeneratedSheet wsNew
            usedNames(sheetName) = True
            UpdateCatalogLink wsDM, r, sheetName
            generatedCount = generatedCount + 1
        End If
    Next r

    wsTemplate.Visible = xlSheetHidden
    UpdateDashboard wsQL, generatedCount, lastRow - 1
    FormatCatalogTable wsDM, lastRow
    ThisWorkbook.Save

CleanExit:
    Application.EnableEvents = True
    Application.DisplayAlerts = True
    Application.ScreenUpdating = True
    Exit Sub

CleanFail:
    MsgBox "Lỗi tạo biên bản: " & Err.Description, vbExclamation, "BienBanNghiemThuTool"
    Resume CleanExit
End Sub

Public Sub GoToDanhMuc()
    ThisWorkbook.Worksheets(SHEET_DM).Activate
End Sub

Public Sub GoToThongTinChung()
    ThisWorkbook.Worksheets(SHEET_DATA).Activate
End Sub

Public Sub SyncBienBan()
    TaoBienBanHangLoat
End Sub

Private Sub DeleteGeneratedSheets()
    Dim i As Long
    Dim nm As String
    For i = ThisWorkbook.Worksheets.Count To 1 Step -1
        nm = ThisWorkbook.Worksheets(i).Name
        If Not IsBaseSheet(nm) Then
            ThisWorkbook.Worksheets(i).Delete
        End If
    Next i
End Sub

Private Function IsBaseSheet(ByVal sheetName As String) As Boolean
    Select Case sheetName
        Case SHEET_QM, SHEET_DM, SHEET_DATA, SHEET_TEMPLATE
            IsBaseSheet = True
        Case Else
            IsBaseSheet = False
    End Select
End Function

Private Function RowHasData(ByVal ws As Worksheet, ByVal r As Long) As Boolean
    Dim c As Long
    For c = 1 To 4
        If Trim$(CStr(ws.Cells(r, c).Value)) <> "" Then
            RowHasData = True
            Exit Function
        End If
    Next c
    RowHasData = False
End Function

Private Function RowSheetBase(ByVal ws As Worksheet, ByVal r As Long) As String
    Dim rawNumber As String
    rawNumber = Trim$(CStr(ws.Cells(r, 1).Value))
    If rawNumber <> "" Then
        RowSheetBase = rawNumber
    Else
        RowSheetBase = "BBNT-" & Format$(r - 1, "000")
    End If
End Function

Private Function UniqueSheetName(ByVal baseName As String, ByVal usedNames As Object) As String
    Dim candidate As String, suffix As Long, suffixText As String
    candidate = CleanSheetName(baseName)
    If Len(candidate) = 0 Then candidate = "BBNT"
    suffix = 1
    Do While usedNames.Exists(candidate) Or SheetExists(candidate)
        suffix = suffix + 1
        suffixText = "-" & Format$(suffix, "00")
        candidate = Left$(CleanSheetName(baseName), 31 - Len(suffixText)) & suffixText
    Loop
    UniqueSheetName = candidate
End Function

Private Function SheetExists(ByVal sheetName As String) As Boolean
    Dim ws As Worksheet
    On Error Resume Next
    Set ws = ThisWorkbook.Worksheets(sheetName)
    SheetExists = Not ws Is Nothing
    On Error GoTo 0
End Function

Private Function CleanSheetName(ByVal textValue As String) As String
    Dim s As String
    s = Trim$(textValue)
    s = Replace(s, "/", "_")
    s = Replace(s, "\", "_")
    s = Replace(s, ":", "_")
    s = Replace(s, "*", "_")
    s = Replace(s, "?", "_")
    s = Replace(s, "[", "_")
    s = Replace(s, "]", "_")
    If Len(s) > 31 Then s = Left$(s, 31)
    CleanSheetName = s
End Function

Private Sub FillAcceptanceSheet(ByVal ws As Worksheet, ByVal wsData As Worksheet, ByVal wsDM As Worksheet, ByVal r As Long)
    Dim docNo As String, workType As String, structureName As String, rawDate As Variant, vnDate As String

    docNo = Trim$(CStr(wsDM.Cells(r, 1).Value))
    If docNo = "" Then docNo = "BBNT-" & Format$(r - 1, "000")
    workType = Trim$(CStr(wsDM.Cells(r, 2).Value))
    structureName = Trim$(CStr(wsDM.Cells(r, 3).Value))
    rawDate = wsDM.Cells(r, 4).Value
    vnDate = FormatVnDate(rawDate)

    ws.Range("A2").Value = "Số / No.: " & docNo
    ws.Range("A6").Value = "Dự án          : " & NzText(wsData.Range("C4").Value)
    ws.Range("A7").Value = "Địa điểm     : " & NzText(wsData.Range("C5").Value)
    ws.Range("C8").Value = ": " & structureName
    ws.Range("C11").Value = structureName
    ws.Range("C14").Value = workType
    ws.Range("D17").Value = ": " & NzText(wsData.Range("C6").Value)
    ws.Range("B18").Value = "Ông : " & NzText(wsData.Range("C7").Value)
    ws.Range("D19").Value = ": " & NzText(wsData.Range("C9").Value)
    ws.Range("D23").Value = ": " & NzText(wsData.Range("C11").Value)
    ws.Range("B24").Value = "Ông : " & NzText(wsData.Range("C12").Value)

    If vnDate <> "" Then
        ws.Range("A28").Value = "  Bắt đầu : 15h00' ngày " & vnDate
        ws.Range("A29").Value = "  Kết thúc : 16h30' ngày " & vnDate
    End If
End Sub

Private Sub FormatGeneratedSheet(ByVal ws As Worksheet)
    On Error Resume Next
    ws.Cells.Font.Name = "Times New Roman"
    ws.Cells.Font.Size = 11
    ws.Rows(1).RowHeight = 6
    ws.Rows(2).Font.Bold = True
    ws.Rows(3).Font.Bold = True
    ws.Rows(4).Font.Bold = True
    ws.Range("A1:G83").WrapText = True
    ws.Columns("A:G").AutoFit
    ws.PageSetup.Orientation = xlPortrait
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = 1
    ws.DisplayGridlines = False
    On Error GoTo 0
End Sub

Private Function NzText(ByVal value As Variant) As String
    If IsError(value) Then
        NzText = ""
    ElseIf Trim$(CStr(value)) = "" Then
        NzText = ""
    Else
        NzText = CStr(value)
    End If
End Function

Private Function FormatVnDate(ByVal value As Variant) As String
    Dim d As Date, parts() As String
    Dim s As String

    If IsDate(value) Then
        d = CDate(value)
        FormatVnDate = Format$(d, "dd") & " tháng " & Format$(d, "mm") & " năm " & Format$(d, "yyyy")
        Exit Function
    End If

    s = Trim$(CStr(value))
    If s = "" Then Exit Function

    If InStr(s, ".") > 0 Then
        parts = Split(s, ".")
    ElseIf InStr(s, "/") > 0 Then
        parts = Split(s, "/")
    ElseIf InStr(s, "-") > 0 Then
        parts = Split(s, "-")
    Else
        FormatVnDate = s
        Exit Function
    End If

    If UBound(parts) = 2 Then
        On Error GoTo FailParse
        d = DateSerial(CInt(parts(2)), CInt(parts(1)), CInt(parts(0)))
        FormatVnDate = Format$(d, "dd") & " tháng " & Format$(d, "mm") & " năm " & Format$(d, "yyyy")
        Exit Function
    End If

FailParse:
    FormatVnDate = s
End Function

Private Sub UpdateCatalogLink(ByVal wsDM As Worksheet, ByVal r As Long, ByVal sheetName As String)
    With wsDM.Cells(r, 5)
        .Value = "Mở"
        On Error Resume Next
        .Hyperlinks.Delete
        On Error GoTo 0
        wsDM.Hyperlinks.Add Anchor:=wsDM.Cells(r, 5), Address:="", SubAddress:="'" & sheetName & "'!A1", TextToDisplay:="Mở"
    End With
End Sub

Private Sub PrepareDashboard(ByVal ws As Worksheet)
    With ws
        .Cells.Clear
        .DisplayGridlines = False
        .Tab.Color = RGB(0, 99, 117)
        .Range("A1:H1").Merge
        .Range("A1").Value = "Bảng quản lý biên bản nghiệm thu"
        .Range("A1").Font.Size = 18
        .Range("A1").Font.Bold = True
        .Range("A1").Font.Color = RGB(255, 255, 255)
        .Range("A1").Interior.Color = RGB(15, 23, 42)
        .Range("A3").Value = "Số dòng danh mục"
        .Range("B3").Formula = "=COUNTA(DanhMuc!A2:A1000)"
        .Range("A4").Value = "Số sheet nghiệm thu"
        .Range("B4").Formula = "=COUNTA(DanhMuc!E2:E1000)"
        .Range("A6").Value = "1. Điền dữ liệu trong sheet DanhMuc."
        .Range("A7").Value = "2. Chỉnh thông tin dùng chung trong sheet ThongTinChung."
        .Range("A8").Value = "3. Bấm nút Đồng bộ để sinh toàn bộ sheet nghiệm thu."
        .Range("A10").Value = "Trạng thái"
        .Range("B10").Value = "Sẵn sàng"
        .Columns("A:H").ColumnWidth = 18
        .Columns("A").ColumnWidth = 30
        .Columns("B").ColumnWidth = 24
        .Rows("3:10").RowHeight = 22
        .Range("A3:B4").Font.Bold = True
        .Range("A3:B4").Interior.Color = RGB(226, 232, 240)
        .Range("A6:B8").WrapText = True
        .Range("A3:B4").Borders.LineStyle = 1
        .Range("A6:B10").Borders.LineStyle = 1
    End With
End Sub

Private Sub PrepareCatalog(ByVal ws As Worksheet)
    With ws
        .DisplayGridlines = False
        .Tab.Color = RGB(34, 197, 94)
        .Cells.Font.Name = "Calibri"
        .Cells.Font.Size = 11
        .Range("A1:E1").Font.Bold = True
        .Range("A1:E1").Interior.Color = RGB(15, 118, 110)
        .Range("A1:E1").Font.Color = RGB(255, 255, 255)
        .Columns("A").ColumnWidth = 18
        .Columns("B").ColumnWidth = 24
        .Columns("C").ColumnWidth = 36
        .Columns("D").ColumnWidth = 18
        .Columns("E").ColumnWidth = 14
        .Rows(1).RowHeight = 22
        .Activate
        .Range("A2").Select
        ActiveWindow.FreezePanes = True
    End With
End Sub

Private Sub PrepareDataSheet(ByVal ws As Worksheet)
    With ws
        .DisplayGridlines = False
        .Tab.Color = RGB(249, 115, 22)
        .Cells.Font.Name = "Calibri"
        .Cells.Font.Size = 11
        .Range("B4:C12").Borders.LineStyle = 1
        .Columns("B").ColumnWidth = 28
        .Columns("C").ColumnWidth = 70
        .Rows("4:12").RowHeight = 22
    End With
End Sub

Private Sub EnsureActionButtons(ByVal ws As Worksheet)
    Dim shp As Shape
    On Error Resume Next
    ws.Shapes("btnSync").Delete
    ws.Shapes("btnDanhMuc").Delete
    ws.Shapes("btnThongTin").Delete
    On Error GoTo 0

    Set shp = ws.Shapes.AddShape(msoShapeRoundedRectangle, 20, 170, 180, 34)
    With shp
        .Name = "btnSync"
        .TextFrame2.TextRange.Text = "Đồng bộ biên bản"
        .OnAction = "SyncBienBan"
        .Fill.ForeColor.RGB = RGB(34, 197, 94)
        .Line.ForeColor.RGB = RGB(22, 163, 74)
        .TextFrame2.TextRange.Font.Fill.ForeColor.RGB = RGB(255, 255, 255)
        .TextFrame2.TextRange.Font.Bold = msoTrue
    End With

    Set shp = ws.Shapes.AddShape(msoShapeRoundedRectangle, 220, 170, 180, 34)
    With shp
        .Name = "btnDanhMuc"
        .TextFrame2.TextRange.Text = "Mở DanhMục"
        .OnAction = "GoToDanhMuc"
        .Fill.ForeColor.RGB = RGB(14, 165, 233)
        .Line.ForeColor.RGB = RGB(2, 132, 199)
        .TextFrame2.TextRange.Font.Fill.ForeColor.RGB = RGB(255, 255, 255)
        .TextFrame2.TextRange.Font.Bold = msoTrue
    End With

    Set shp = ws.Shapes.AddShape(msoShapeRoundedRectangle, 420, 170, 180, 34)
    With shp
        .Name = "btnThongTin"
        .TextFrame2.TextRange.Text = "Mở ThongTinChung"
        .OnAction = "GoToThongTinChung"
        .Fill.ForeColor.RGB = RGB(249, 115, 22)
        .Line.ForeColor.RGB = RGB(234, 88, 12)
        .TextFrame2.TextRange.Font.Fill.ForeColor.RGB = RGB(255, 255, 255)
        .TextFrame2.TextRange.Font.Bold = msoTrue
    End With
End Sub

Private Sub FormatCatalogTable(ByVal ws As Worksheet, ByVal lastRow As Long)
    Dim lo As ListObject
    On Error Resume Next
    For Each lo In ws.ListObjects
        lo.Unlist
    Next lo
    On Error GoTo 0
    If lastRow < 2 Then Exit Sub
    Set lo = ws.ListObjects.Add(xlSrcRange, ws.Range("A1:E" & lastRow), , xlYes)
    lo.Name = "tblDanhMuc"
    lo.TableStyle = "TableStyleMedium2"
End Sub

Private Sub UpdateDashboard(ByVal ws As Worksheet, ByVal generatedCount As Long, ByVal catalogRows As Long)
    With ws
        .Range("B10").Value = "Số sheet: " & generatedCount & " / Dòng danh mục: " & catalogRows
        .Range("B10").Font.Bold = True
        .Range("B10").Font.Color = RGB(15, 23, 42)
    End With
End Sub
'''


THISWORKBOOK_CODE = r'''
Option Explicit

Private Sub Workbook_SheetChange(ByVal Sh As Object, ByVal Target As Range)
    On Error GoTo SafeExit
    If Sh.Name = "DanhMuc" Then
        If Not Intersect(Target, Sh.Range("A2:E1000")) Is Nothing Then
            Application.EnableEvents = False
            TaoBienBanHangLoat
        End If
    End If
SafeExit:
    Application.EnableEvents = True
End Sub
'''


def delete_if_exists(path: Path) -> None:
    try:
        if path.exists():
            path.unlink()
    except PermissionError:
        raise RuntimeError(f"Không thể ghi đè file đang mở: {path}")


def prepare_dashboard_com(ws) -> None:
    ws.Cells.Clear()
    try:
        ws.Application.ActiveWindow.DisplayGridlines = False
    except Exception:
        pass
    ws.Tab.Color = 0x007063
    ws.Range("A1:H1").Merge()
    ws.Range("A1").Value = "Bảng quản lý biên bản nghiệm thu"
    ws.Range("A1").Font.Size = 18
    ws.Range("A1").Font.Bold = True
    ws.Range("A1").Font.Color = 0xFFFFFF
    ws.Range("A1").Interior.Color = 0x1A172A
    ws.Range("A3").Value = "Số dòng danh mục"
    ws.Range("B3").Formula = "=COUNTA(DanhMuc!A2:A1000)"
    ws.Range("A4").Value = "Số sheet nghiệm thu"
    ws.Range("B4").Formula = "=COUNTA(DanhMuc!E2:E1000)"
    ws.Range("A6").Value = "1. Điền dữ liệu trong sheet DanhMuc."
    ws.Range("A7").Value = "2. Chỉnh thông tin dùng chung trong sheet ThongTinChung."
    ws.Range("A8").Value = "3. Bấm nút Đồng bộ để sinh toàn bộ sheet nghiệm thu."
    ws.Range("A10").Value = "Trạng thái"
    ws.Range("B10").Value = "Sẵn sàng"
    ws.Columns("A").ColumnWidth = 30
    ws.Columns("B").ColumnWidth = 24
    ws.Columns("C:H").ColumnWidth = 18
    ws.Range("A3:B4").Font.Bold = True
    ws.Range("A3:B4").Interior.Color = 0xE8E2E2
    ws.Range("A6:B10").WrapText = True
    ws.Range("A3:B4").Borders.LineStyle = 1
    ws.Range("A6:B10").Borders.LineStyle = 1


def prepare_data_com(ws, project: ProjectData) -> None:
    try:
        ws.Application.ActiveWindow.DisplayGridlines = False
    except Exception:
        pass
    ws.Tab.Color = 0x1E8BFA
    ws.Cells.Font.Name = "Calibri"
    ws.Cells.Font.Size = 11
    ws.Range("B4:C12").Borders.LineStyle = 1
    ws.Columns("B").ColumnWidth = 28
    ws.Columns("C").ColumnWidth = 70
    ws.Rows("4:12").RowHeight = 22
    ws.Range("C4").Value = project.project
    ws.Range("C5").Value = project.location
    ws.Range("C6").Value = project.owner
    ws.Range("C7").Value = project.owner_rep
    ws.Range("C8").Value = " " if not project.consultant else ""
    ws.Range("C9").Value = project.consultant
    ws.Range("C11").Value = project.contractor
    ws.Range("C12").Value = project.contractor_rep


def prepare_catalog_com(ws) -> None:
    try:
        ws.Application.ActiveWindow.DisplayGridlines = False
    except Exception:
        pass
    ws.Tab.Color = 0x22C55E
    ws.Cells.Font.Name = "Calibri"
    ws.Cells.Font.Size = 11
    ws.Range("A1:E1").Font.Bold = True
    ws.Range("A1:E1").Interior.Color = 0x0F767E
    ws.Range("A1:E1").Font.Color = 0xFFFFFF
    ws.Columns("A").ColumnWidth = 18
    ws.Columns("B").ColumnWidth = 24
    ws.Columns("C").ColumnWidth = 36
    ws.Columns("D").ColumnWidth = 18
    ws.Columns("E").ColumnWidth = 14
    ws.Rows(1).RowHeight = 22


def add_action_buttons_com(ws) -> None:
    for shape_name in ("btnSync", "btnDanhMuc", "btnThongTin"):
        try:
            ws.Shapes(shape_name).Delete()
        except Exception:
            pass

    buttons = [
        ("btnSync", "Đồng bộ biên bản", "SyncBienBan", 20, 170, 180, 34, 0x22C55E, 0x16A34A),
        ("btnDanhMuc", "Mở DanhMục", "GoToDanhMuc", 220, 170, 180, 34, 0x0EA5E9, 0x0284C7),
        ("btnThongTin", "Mở ThongTinChung", "GoToThongTinChung", 420, 170, 180, 34, 0xF97316, 0xEA580C),
    ]
    for name, text, action, left, top, width, height, fill, line in buttons:
        shp = ws.Shapes.AddShape(5, left, top, width, height)
        shp.Name = name
        shp.TextFrame2.TextRange.Text = text
        shp.OnAction = action
        shp.Fill.ForeColor.RGB = fill
        shp.Line.ForeColor.RGB = line
        shp.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0xFFFFFF
        shp.TextFrame2.TextRange.Font.Bold = True


def add_dashboard_hyperlinks(ws) -> None:
    ws.Hyperlinks.Add(Anchor=ws.Range("B10"), Address="", SubAddress="'DanhMuc'!A1", TextToDisplay="Mở")
    ws.Hyperlinks.Add(Anchor=ws.Range("B11"), Address="", SubAddress="'ThongTinChung'!A1", TextToDisplay="Mở")


def clear_dynamic_fields_com(ws) -> None:
    for addr in ("A2", "A6", "A7", "C8", "C11", "C14", "D17", "B18", "D19", "B20", "D20", "B21", "D21", "B22", "D22", "D23", "B24", "A28", "A29"):
        ws.Range(addr).Value = ""


def apply_row_com(ws, row: BatchRow, project: ProjectData) -> None:
    ws.Range("A2").Value = f"Số / No.: {row.doc_number}"
    ws.Range("A6").Value = f"Dự án          : {project.project}"
    ws.Range("A7").Value = f"Địa điểm     : {project.location}"
    ws.Range("C8").Value = f": {row.structure_name}"
    ws.Range("C11").Value = row.structure_name
    ws.Range("C14").Value = row.work_type
    ws.Range("D17").Value = f": {project.owner}"
    ws.Range("B18").Value = f"Ông : {project.owner_rep}"
    ws.Range("D19").Value = f": {project.consultant}"
    ws.Range("D23").Value = f": {project.contractor}"
    ws.Range("B24").Value = f"Ông : {project.contractor_rep}"
    vn_date = format_vn_date(row.date_text)
    if vn_date:
        ws.Range("A28").Value = f"  Bắt đầu : 15h00' ngày {vn_date}"
        ws.Range("A29").Value = f"  Kết thúc : 16h30' ngày {vn_date}"


def format_generated_sheet_com(ws) -> None:
    ws.Cells.Font.Name = "Times New Roman"
    ws.Cells.Font.Size = 11
    ws.Rows(1).RowHeight = 6
    try:
        ws.Application.ActiveWindow.DisplayGridlines = False
    except Exception:
        pass
    ws.Range("A1:G83").WrapText = True
    ws.Columns("A:G").AutoFit()


def set_catalog_row_com(ws, row_num: int, row: BatchRow) -> None:
    ws.Cells(row_num, 1).Value = row.doc_number
    ws.Cells(row_num, 2).Value = row.work_type
    ws.Cells(row_num, 3).Value = row.structure_name
    ws.Cells(row_num, 4).Value = row.date_text
    ws.Cells(row_num, 5).Value = "Mở"


def set_summary_row_com(ws, row_num: int, row: BatchRow) -> None:
    ws.Cells(row_num, 1).Value = row_num - 1
    ws.Cells(row_num, 2).Value = row.doc_number
    ws.Cells(row_num, 3).Value = row.work_type
    ws.Cells(row_num, 4).Value = row.structure_name
    ws.Cells(row_num, 5).Value = row.date_text
    ws.Cells(row_num, 6).Value = "Mở"


def create_catalog_table_com(ws, last_row: int) -> None:
    try:
        for lo in list(ws.ListObjects):
            lo.Unlist()
    except Exception:
        pass
    if last_row >= 2:
        lo = ws.ListObjects.Add(1, ws.Range(f"A1:E{last_row}"), None, 1)
        lo.Name = "tblDanhMuc"
        lo.TableStyle = "TableStyleMedium2"


def build_initial_workbook_state(wb, batch_rows: list[BatchRow], project: ProjectData) -> None:
    ws_qm = wb.Worksheets("QuanLy")
    ws_data = wb.Worksheets("ThongTinChung")
    ws_dm = wb.Worksheets("DanhMuc")
    ws_template = wb.Worksheets("Mau_NT")

    prepare_dashboard_com(ws_qm)
    prepare_data_com(ws_data, project)
    prepare_catalog_com(ws_dm)
    add_action_buttons_com(ws_qm)
    add_dashboard_hyperlinks(ws_qm)

    # Clear any pre-existing generated sheets just in case.
    for i in range(wb.Worksheets.Count, 0, -1):
        name = wb.Worksheets(i).Name
        if name not in {"QuanLy", "ThongTinChung", "DanhMuc", "Mau_NT"}:
            wb.Worksheets(i).Delete()

    used_names = {"QuanLy", "ThongTinChung", "DanhMuc", "Mau_NT"}
    for idx, row in enumerate(batch_rows, start=2):
        ws_template.Copy(None, wb.Worksheets(wb.Worksheets.Count))
        ws_new = wb.Worksheets(wb.Worksheets.Count)
        clear_dynamic_fields_com(ws_new)
        apply_row_com(ws_new, row, project)
        name = safe_sheet_name(row.doc_number, used_names)
        ws_new.Name = name
        used_names.add(name)
        row.link = name
        format_generated_sheet_com(ws_new)

    # Rebuild catalog with links and dashboard counters.
    ws_dm.Cells.ClearContents()
    ws_dm.Range("A1:E1").Value = [["SỐ BBNT", "CÔNG TÁC NT", "TÊN CẤU KIỆN", "NGÀY NGHIỆM THU", "LINK"]]
    for idx, row in enumerate(batch_rows, start=2):
        set_catalog_row_com(ws_dm, idx, row)
        ws_dm.Hyperlinks.Add(Anchor=ws_dm.Cells(idx, 5), Address="", SubAddress=f"'{row.link}'!A1", TextToDisplay="Mở")

    ws_qm.Range("B10").Value = f"Số sheet: {len(batch_rows)} / Dòng danh mục: {len(batch_rows)}"
    create_catalog_table_com(ws_dm, len(batch_rows) + 1)


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not SOURCE_XLS.exists():
        raise FileNotFoundError(SOURCE_XLS)

    OUTPUT_XLSM.parent.mkdir(parents=True, exist_ok=True)
    delete_if_exists(OUTPUT_XLSM)
    delete_if_exists(DESKTOP_COPY)
    temp_source_xlsx = convert_source_to_xlsx(SOURCE_XLS)
    batch_rows, project = load_source_data(str(temp_source_xlsx))

    excel = Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    wb = None
    src_wb = None
    try:
        src_wb = excel.Workbooks.Open(str(SOURCE_XLS), ReadOnly=True)
        src_wb.SaveAs(str(OUTPUT_XLSM), FileFormat=52)
        src_wb.Close(False)
        src_wb = None

        wb = excel.Workbooks.Open(str(OUTPUT_XLSM), ReadOnly=False)

        # Base sheets become the workbook shell.
        ws_data = wb.Worksheets("Data")
        ws_data.Name = "ThongTinChung"
        ws_dm = wb.Worksheets("DANH MUC NT TSX")
        ws_dm.Name = "DanhMuc"
        ws_template = wb.Worksheets("NT NEN DAT MONG TSX 17.06 ")
        ws_template.Name = "Mau_NT"

        for i in range(wb.Worksheets.Count, 0, -1):
            name = wb.Worksheets(i).Name
            if name not in {"ThongTinChung", "DanhMuc", "Mau_NT"}:
                wb.Worksheets(i).Delete()

        ws_qm = wb.Worksheets.Add(Before=wb.Worksheets(1))
        ws_qm.Name = "QuanLy"
        build_initial_workbook_state(wb, batch_rows, project)
        ws_template.Visible = 0
        wb.Save()

        wb.Close(SaveChanges=False)
        wb = excel.Workbooks.Open(str(OUTPUT_XLSM), ReadOnly=False)

        try:
            vbproj = wb.VBProject
        except Exception as exc:
            raise RuntimeError("Excel không cho phép truy cập VBA project. Hãy bật 'Trust access to the VBA project object model'.") from exc

        for idx in range(vbproj.VBComponents.Count, 0, -1):
            comp = vbproj.VBComponents.Item(idx)
            if comp.Type == 1 and comp.Name not in {"ThisWorkbook"}:
                try:
                    vbproj.VBComponents.Remove(comp)
                except Exception:
                    pass

        mod = vbproj.VBComponents.Add(1)
        mod.Name = "modBienBan"
        mod.CodeModule.AddFromString(VBA_MODULE)

        thiswb = vbproj.VBComponents("ThisWorkbook")
        if thiswb.CodeModule.CountOfLines > 0:
            thiswb.CodeModule.DeleteLines(1, thiswb.CodeModule.CountOfLines)
        thiswb.CodeModule.AddFromString(THISWORKBOOK_CODE)

        wb.Save()

        shutil.copy2(OUTPUT_XLSM, DESKTOP_COPY)
        print(f"Đã tạo workbook: {OUTPUT_XLSM}")
        print(f"Đã sao chép ra Desktop: {DESKTOP_COPY}")
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        if src_wb is not None:
            src_wb.Close(False)
        excel.Quit()
        try:
            shutil.rmtree(temp_source_xlsx.parent, ignore_errors=True)
        except Exception:
            pass


if __name__ == "__main__":
    main()
